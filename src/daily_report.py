#!/usr/bin/env python3
"""
Eachlabs daily execution report builder.

What it does:
1. Picks the newest execution_report_*.{xlsx,csv} from ./inbox
2. Reads the raw data (xlsx sheet 0 or csv)
3. Builds an aggregated org x date spend table
4. Computes top movers (today vs yesterday), first-time spenders today, drop-offs
5. Writes a fully formatted xlsx into ./reports
6. Moves the source file into ./archive

Run:
    python3 daily_report.py /path/to/eachlabs-daily

If the path arg is omitted, defaults to the parent of this script's directory.
"""

from __future__ import annotations

import os
import sys
import shutil
import json
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

SPEND_COL = "total_revenue"  # what customers spend (= Eachlabs revenue)
DROP_OFF_LOOKBACK_DAYS = 7
DROP_OFF_MIN_SPEND = 5.0  # only flag drop-offs that spent at least this in lookback
TOP_MOVERS_LIMIT = 25
FIRST_TIME_LIMIT = 50
MONTHLY_PIVOT_MONTHS = 6
MONTHLY_COCKPIT_MONTHS = 6
MODEL_WEEKLY_WEEKS = 8
MODEL_PIVOT_TOP_N = 20
CUSTOMER_TIERS = [10000, 5000, 1000, 500]  # dollar bands for tier counts


SUPPORTED_EXTS = (".xlsx", ".csv")


def find_newest_inbox_file(inbox: Path) -> Path | None:
    candidates = []
    for ext in SUPPORTED_EXTS:
        candidates.extend(inbox.glob(f"execution_report_*{ext}"))
    files = sorted(
        [p for p in candidates if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


def load_raw(src_path: Path) -> pd.DataFrame:
    # Accept .xlsx (sheet 0) or .csv. CSVs are written by Slack exports / BI tools
    # and don't need the openpyxl round-trip.
    suffix = src_path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(src_path)
    elif suffix in (".xlsx", ".xlsm"):
        df = pd.read_excel(src_path, sheet_name=0, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported file type: {src_path.suffix} ({src_path.name})")
    df["execution_date"] = pd.to_datetime(df["execution_date"]).dt.normalize()
    df = df.dropna(subset=["organization_name", "execution_date"])
    df["organization_name"] = df["organization_name"].astype(str).str.strip()
    for col in ["execution_count", "total_revenue", "total_provider_cost", "total_margin"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            # Upstream export occasionally lags on cost/margin sync.
            # Default to 0 rather than crash the whole pipeline.
            df[col] = 0.0
    return df


def org_daily_spend(df: pd.DataFrame) -> pd.DataFrame:
    g = (
        df.groupby(["organization_name", "execution_date"], as_index=False)[SPEND_COL]
        .sum()
    )
    return g


MTD_WINDOW_DAYS = 30  # rolling window — "MTD" here = trailing 30 days


def build_mtd_pivot(daily: pd.DataFrame) -> pd.DataFrame:
    """Pivot of org x day for the trailing 30-day window."""
    if daily.empty:
        return pd.DataFrame()
    max_date = daily["execution_date"].max()
    window_start = max_date - pd.Timedelta(days=MTD_WINDOW_DAYS - 1)
    recent = daily[
        (daily["execution_date"] >= window_start)
        & (daily["execution_date"] <= max_date)
    ].copy()
    if recent.empty:
        return pd.DataFrame()
    pivot = recent.pivot_table(
        index="organization_name",
        columns="execution_date",
        values=SPEND_COL,
        aggfunc="sum",
        fill_value=0,
    )
    # ensure every day in the window is present even if no spend
    full_days = pd.date_range(window_start, max_date, freq="D")
    pivot = pivot.reindex(columns=full_days, fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    # drop orgs with zero spend in the window so the sheet stays focused
    pivot = pivot[pivot["Total"] > 0]
    return pivot


def build_monthly_pivot(daily: pd.DataFrame, months: int) -> pd.DataFrame:
    if daily.empty:
        return pd.DataFrame()
    d = daily.copy()
    d["month"] = d["execution_date"].dt.to_period("M").dt.to_timestamp()
    max_month = d["month"].max()
    # keep last `months` months
    cutoff = (max_month - pd.DateOffset(months=months - 1)).to_pydatetime()
    d = d[d["month"] >= cutoff]
    pivot = d.pivot_table(
        index="organization_name",
        columns="month",
        values=SPEND_COL,
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    return pivot


def build_monthly_cockpit(raw: pd.DataFrame, months: int = MONTHLY_COCKPIT_MONTHS) -> pd.DataFrame:
    """Transposed monthly view: metrics on rows, months on columns.
    Returns a DataFrame where the index is the metric name and columns are months."""
    if raw.empty:
        return pd.DataFrame()
    d = raw.copy()
    d["month"] = d["execution_date"].dt.to_period("M")
    keep_months = sorted(d["month"].unique())[-months:]
    d = d[d["month"].isin(keep_months)]

    g = d.groupby("month").agg(
        Revenue=("total_revenue", "sum"),
        Cost=("total_provider_cost", "sum"),
        Margin=("total_margin", "sum"),
        Executions=("execution_count", "sum"),
    ).reset_index()
    g["Margin %"] = (g["Margin"] / g["Revenue"] * 100).round(1)
    active = d.groupby("month")["organization_name"].nunique()
    g["Active orgs"] = g["month"].map(active).fillna(0).astype(int)
    g["Avg $/org"] = (g["Revenue"] / g["Active orgs"].replace(0, np.nan)).round(0)
    g["MoM growth %"] = (g["Revenue"].pct_change() * 100).round(1)

    # Customer tier counts — orgs that spent above each tier in that month
    org_month = d.groupby(["month", "organization_name"])["total_revenue"].sum().reset_index()
    for t in CUSTOMER_TIERS:
        label = f"# orgs > ${t/1000:.1f}k" if t < 1000 else f"# orgs > ${t//1000}k"
        counts = org_month[org_month["total_revenue"] > t].groupby("month").size()
        g[label] = g["month"].map(counts).fillna(0).astype(int)

    g["month"] = g["month"].astype(str)
    return g.set_index("month").T


def build_monthly_model_winners(raw: pd.DataFrame, months: int = MONTHLY_COCKPIT_MONTHS) -> pd.DataFrame:
    """Per-month top model by executions and by revenue. Last N months."""
    if raw.empty:
        return pd.DataFrame()
    d = raw.copy()
    d["month"] = d["execution_date"].dt.to_period("M")
    keep = sorted(d["month"].unique())[-months:]
    d = d[d["month"].isin(keep)]
    agg = d.groupby(["month", "model_name"]).agg(
        executions=("execution_count", "sum"),
        revenue=("total_revenue", "sum"),
    ).reset_index()
    top_e = agg.sort_values(["month", "executions"], ascending=[True, False]).groupby("month").head(1)
    top_r = agg.sort_values(["month", "revenue"], ascending=[True, False]).groupby("month").head(1)
    out = top_e.merge(top_r, on="month", suffixes=("_by_exec", "_by_rev"))
    out["month"] = out["month"].astype(str)
    return out[[
        "month",
        "model_name_by_exec", "executions_by_exec",
        "model_name_by_rev", "revenue_by_rev",
    ]].rename(columns={
        "model_name_by_exec": "top model (executions)",
        "executions_by_exec": "executions",
        "model_name_by_rev": "top model (revenue)",
        "revenue_by_rev": "revenue",
    })


SEEDANCE_2_REGEX = r"seedance-2[\.\-]0"  # matches seedance-2.0 and seedance-2-0
SEEDANCE_EXCLUDE_REGEX = r"emre-test"  # dev/test models to exclude


def build_seedance_usage(raw: pd.DataFrame) -> tuple[dict, pd.DataFrame, pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    """Filter to Seedance 2.0 models and return:
      - summary dict (totals + unique orgs)
      - daily breakdown (date, executions, revenue, orgs)
      - workspaces x days pivot (org rows, date cols, execution counts)
      - window_start, window_end
    """
    if raw.empty:
        empty_df = pd.DataFrame()
        return {}, empty_df, empty_df, pd.NaT, pd.NaT

    mask = (
        raw["model_name"].str.contains(SEEDANCE_2_REGEX, case=False, na=False, regex=True)
        & ~raw["model_name"].str.contains(SEEDANCE_EXCLUDE_REGEX, case=False, na=False, regex=True)
    )
    sd = raw[mask].copy()

    latest = raw["execution_date"].max()
    window_start = latest - pd.Timedelta(days=MTD_WINDOW_DAYS - 1)
    sd_win = sd[(sd["execution_date"] >= window_start) & (sd["execution_date"] <= latest)]

    if sd_win.empty:
        summary = {
            "window_start": window_start,
            "window_end": latest,
            "total_executions": 0,
            "total_revenue": 0.0,
            "total_cost": 0.0,
            "total_margin": 0.0,
            "margin_pct": 0.0,
            "unique_orgs": 0,
            "active_days": 0,
            "days_in_window": MTD_WINDOW_DAYS,
        }
        return summary, pd.DataFrame(), pd.DataFrame(), window_start, latest

    total_rev = sd_win["total_revenue"].sum()
    summary = {
        "window_start": window_start,
        "window_end": latest,
        "total_executions": int(sd_win["execution_count"].sum()),
        "total_revenue": float(total_rev),
        "total_cost": float(sd_win["total_provider_cost"].sum()),
        "total_margin": float(sd_win["total_margin"].sum()),
        "margin_pct": float(sd_win["total_margin"].sum() / total_rev * 100) if total_rev else 0.0,
        "unique_orgs": int(sd_win["organization_name"].nunique()),
        "active_days": int(sd_win["execution_date"].nunique()),
        "days_in_window": MTD_WINDOW_DAYS,
    }

    # Daily breakdown (newest first)
    daily = sd_win.groupby("execution_date").agg(
        executions=("execution_count", "sum"),
        revenue=("total_revenue", "sum"),
        active_orgs=("organization_name", "nunique"),
    ).reset_index().sort_values("execution_date", ascending=False)

    # Workspaces x days pivot
    pivot = sd_win.pivot_table(
        index="organization_name",
        columns="execution_date",
        values="execution_count",
        aggfunc="sum",
        fill_value=0,
    )
    full_days = pd.date_range(window_start, latest, freq="D")
    pivot = pivot.reindex(columns=full_days, fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)

    return summary, daily, pivot, window_start, latest


def build_top_models_30d(raw: pd.DataFrame, top_n: int = MODEL_PIVOT_TOP_N) -> tuple[pd.DataFrame, pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    """Top N models in the trailing 30-day window, sorted by executions and by revenue.
    Returns (by_executions_df, by_revenue_df, window_start, window_end)."""
    if raw.empty:
        empty = pd.DataFrame(columns=["#", "model", "executions", "revenue", "margin %"])
        return empty, empty, pd.NaT, pd.NaT
    latest = raw["execution_date"].max()
    window_start = latest - pd.Timedelta(days=MTD_WINDOW_DAYS - 1)
    win = raw[(raw["execution_date"] >= window_start) & (raw["execution_date"] <= latest)]
    agg = win.groupby("model_name").agg(
        executions=("execution_count", "sum"),
        revenue=("total_revenue", "sum"),
        cost=("total_provider_cost", "sum"),
        margin=("total_margin", "sum"),
    ).reset_index()
    agg["margin %"] = (agg["margin"] / agg["revenue"] * 100).where(agg["revenue"] > 0).round(1)

    by_exec = agg.sort_values("executions", ascending=False).head(top_n).reset_index(drop=True)
    by_exec.insert(0, "#", range(1, len(by_exec) + 1))
    by_exec = by_exec[["#", "model_name", "executions", "revenue", "margin %"]].rename(columns={"model_name": "model"})

    by_rev = agg.sort_values("revenue", ascending=False).head(top_n).reset_index(drop=True)
    by_rev.insert(0, "#", range(1, len(by_rev) + 1))
    by_rev = by_rev[["#", "model_name", "revenue", "executions", "margin %"]].rename(columns={"model_name": "model"})

    return by_exec, by_rev, window_start, latest


def build_weekly_winners(raw: pd.DataFrame, weeks: int = MODEL_WEEKLY_WEEKS) -> pd.DataFrame:
    """Per-week top model by executions and by revenue."""
    if raw.empty:
        return pd.DataFrame()
    d = raw.copy()
    d["week"] = d["execution_date"].dt.to_period("W-SUN")
    keep_weeks = sorted(d["week"].unique())[-weeks:]
    d = d[d["week"].isin(keep_weeks)]
    agg = d.groupby(["week", "model_name"]).agg(
        executions=("execution_count", "sum"),
        revenue=("total_revenue", "sum"),
    ).reset_index()
    top_exec = agg.sort_values(["week", "executions"], ascending=[True, False]).groupby("week").head(1)
    top_rev = agg.sort_values(["week", "revenue"], ascending=[True, False]).groupby("week").head(1)
    out = top_exec.merge(top_rev, on="week", suffixes=("_by_exec", "_by_rev"))
    out["week"] = out["week"].astype(str)
    return out[[
        "week",
        "model_name_by_exec", "executions_by_exec",
        "model_name_by_rev", "revenue_by_rev",
    ]].rename(columns={
        "model_name_by_exec": "top model (executions)",
        "executions_by_exec": "executions",
        "model_name_by_rev": "top model (revenue)",
        "revenue_by_rev": "revenue",
    })


def build_weekly_model_pivot(
    raw: pd.DataFrame,
    weeks: int = MODEL_WEEKLY_WEEKS,
    top_n: int = MODEL_PIVOT_TOP_N,
    value_col: str = "execution_count",
) -> pd.DataFrame:
    """Pivot of model x week. Keeps the top N models by total in the window."""
    if raw.empty:
        return pd.DataFrame()
    d = raw.copy()
    d["week"] = d["execution_date"].dt.to_period("W-SUN")
    keep_weeks = sorted(d["week"].unique())[-weeks:]
    d = d[d["week"].isin(keep_weeks)]
    pivot = d.pivot_table(
        index="model_name",
        columns="week",
        values=value_col,
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False).head(top_n)
    # rename columns to strings for the writer
    pivot.columns = [str(c) for c in pivot.columns]
    return pivot


def build_projection(raw: pd.DataFrame) -> dict:
    """Current calendar month: actuals + linear projection + MoM + tier counts.
    Returns a dict the writer can render directly."""
    if raw.empty:
        return {}
    latest = raw["execution_date"].max()
    cal_start = latest.replace(day=1)
    cal_end = (cal_start + pd.offsets.MonthEnd(0)).normalize()
    days_in_month = (cal_end - cal_start).days + 1
    days_elapsed = (latest - cal_start).days + 1

    cm = raw[(raw["execution_date"] >= cal_start) & (raw["execution_date"] <= latest)]
    rev = cm["total_revenue"].sum()
    cost = cm["total_provider_cost"].sum()
    margin = cm["total_margin"].sum()
    execs = cm["execution_count"].sum()

    daily_rev = rev / days_elapsed if days_elapsed else 0
    daily_cost = cost / days_elapsed if days_elapsed else 0
    daily_margin = margin / days_elapsed if days_elapsed else 0
    proj_rev = daily_rev * days_in_month
    proj_cost = daily_cost * days_in_month
    proj_margin = daily_margin * days_in_month

    # Previous calendar month for MoM
    prev_start = (cal_start - pd.DateOffset(months=1)).normalize()
    prev_end = cal_start - pd.Timedelta(days=1)
    pm = raw[(raw["execution_date"] >= prev_start) & (raw["execution_date"] <= prev_end)]
    pm_rev = pm["total_revenue"].sum()
    pm_margin = pm["total_margin"].sum()

    # Customer tier counts — actual so far, plus projected at month end
    org_now = cm.groupby("organization_name")["total_revenue"].sum()
    org_proj = org_now * (days_in_month / days_elapsed) if days_elapsed else org_now
    tiers = []
    for t in CUSTOMER_TIERS:
        label = f"${t/1000:.1f}k" if t < 1000 else f"${t//1000}k"
        tiers.append({
            "threshold": label,
            "current": int((org_now > t).sum()),
            "projected": int((org_proj > t).sum()),
        })

    return {
        "month_label": cal_start.strftime("%B %Y"),
        "prev_month_label": prev_start.strftime("%b %Y"),
        "days_in_month": days_in_month,
        "days_elapsed": days_elapsed,
        "days_remaining": days_in_month - days_elapsed,
        "through_date": latest.strftime("%Y-%m-%d"),
        "rev_actual": rev,
        "cost_actual": cost,
        "margin_actual": margin,
        "execs_actual": int(execs),
        "daily_avg_rev": daily_rev,
        "rev_proj": proj_rev,
        "cost_proj": proj_cost,
        "margin_proj": proj_margin,
        "prev_rev": pm_rev,
        "prev_margin": pm_margin,
        "mom_rev_pct": ((proj_rev - pm_rev) / pm_rev * 100) if pm_rev else None,
        "mom_margin_pct": ((proj_margin - pm_margin) / pm_margin * 100) if pm_margin else None,
        "tiers": tiers,
    }


def compute_movers(daily: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    if daily.empty:
        empty = pd.DataFrame(columns=["organization_name", "yesterday", "today", "delta", "pct_change"])
        return empty, empty, pd.NaT, pd.NaT
    dates = sorted(daily["execution_date"].unique())
    today = dates[-1]
    yesterday = dates[-2] if len(dates) > 1 else None

    today_df = daily[daily["execution_date"] == today].set_index("organization_name")[SPEND_COL]
    if yesterday is not None:
        yest_df = daily[daily["execution_date"] == yesterday].set_index("organization_name")[SPEND_COL]
    else:
        yest_df = pd.Series(dtype=float)

    combined = pd.DataFrame({"yesterday": yest_df, "today": today_df}).fillna(0)
    combined["delta"] = combined["today"] - combined["yesterday"]
    combined["pct_change"] = np.where(
        combined["yesterday"] > 0,
        (combined["delta"] / combined["yesterday"]) * 100,
        np.nan,
    )
    combined = combined.reset_index()
    movers_up = combined.sort_values("delta", ascending=False).head(TOP_MOVERS_LIMIT)
    movers_down = combined.sort_values("delta", ascending=True).head(TOP_MOVERS_LIMIT)
    return movers_up, movers_down, today, yesterday


def compute_first_time_today(daily: pd.DataFrame) -> pd.DataFrame:
    if daily.empty:
        return pd.DataFrame(columns=["organization_name", "first_spend_date", "today_spend"])
    today = daily["execution_date"].max()
    first_dates = (
        daily[daily[SPEND_COL] > 0]
        .groupby("organization_name", as_index=False)["execution_date"]
        .min()
        .rename(columns={"execution_date": "first_spend_date"})
    )
    first_today = first_dates[first_dates["first_spend_date"] == today]
    today_spend = (
        daily[daily["execution_date"] == today]
        .groupby("organization_name", as_index=False)[SPEND_COL]
        .sum()
        .rename(columns={SPEND_COL: "today_spend"})
    )
    out = first_today.merge(today_spend, on="organization_name", how="left")
    out = out.sort_values("today_spend", ascending=False).head(FIRST_TIME_LIMIT)
    return out


def compute_drop_offs(daily: pd.DataFrame) -> pd.DataFrame:
    if daily.empty:
        return pd.DataFrame(
            columns=["organization_name", "lookback_total", "last_spend_date", "days_since"]
        )
    today = daily["execution_date"].max()
    lookback_start = today - pd.Timedelta(days=DROP_OFF_LOOKBACK_DAYS)
    today_set = set(daily[daily["execution_date"] == today]["organization_name"].unique())
    window = daily[
        (daily["execution_date"] > lookback_start)
        & (daily["execution_date"] < today)
        & (daily[SPEND_COL] > 0)
    ]
    grouped = (
        window.groupby("organization_name")
        .agg(lookback_total=(SPEND_COL, "sum"), last_spend_date=("execution_date", "max"))
        .reset_index()
    )
    grouped = grouped[
        (~grouped["organization_name"].isin(today_set))
        & (grouped["lookback_total"] >= DROP_OFF_MIN_SPEND)
    ]
    grouped["days_since"] = (today - grouped["last_spend_date"]).dt.days
    grouped = grouped.sort_values("lookback_total", ascending=False).head(FIRST_TIME_LIMIT)
    return grouped


def fmt_money(v):
    try:
        return float(v)
    except Exception:
        return v


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
SUBTLE_FILL = PatternFill("solid", fgColor="F3F4F6")
THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, df_cols):
    for i, col in enumerate(df_cols, start=1):
        letter = get_column_letter(i)
        max_len = max(
            [len(str(col))]
            + [
                len(str(ws.cell(row=r, column=i).value)) if ws.cell(row=r, column=i).value is not None else 0
                for r in range(2, ws.max_row + 1)
            ]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)


def write_summary_sheet(ws, today, yesterday, daily, movers_up, movers_down, first_time, drop_offs):
    today_total = daily[daily["execution_date"] == today][SPEND_COL].sum() if not daily.empty else 0
    yest_total = (
        daily[daily["execution_date"] == yesterday][SPEND_COL].sum()
        if (yesterday is not None and not daily.empty)
        else 0
    )
    last7_avg = 0
    if not daily.empty:
        cutoff = today - pd.Timedelta(days=7)
        last7 = daily[(daily["execution_date"] > cutoff) & (daily["execution_date"] < today)]
        if not last7.empty:
            last7_avg = last7.groupby("execution_date")[SPEND_COL].sum().mean()

    mtd_total = 0
    last_month_pace = 0
    if not daily.empty:
        # Rolling 30-day window
        window_start = today - pd.Timedelta(days=MTD_WINDOW_DAYS - 1)
        mtd = daily[(daily["execution_date"] >= window_start) & (daily["execution_date"] <= today)]
        mtd_total = mtd[SPEND_COL].sum()

        # Compare against the prior 30 days (days -60..-30 from today)
        prior_start = window_start - pd.Timedelta(days=MTD_WINDOW_DAYS)
        prior_end = window_start - pd.Timedelta(days=1)
        prior_window = daily[
            (daily["execution_date"] >= prior_start)
            & (daily["execution_date"] <= prior_end)
        ]
        last_month_pace = prior_window[SPEND_COL].sum()

    rows = [
        ("Report date", today.strftime("%Y-%m-%d") if today is not pd.NaT else "n/a"),
        ("Today total spend", today_total),
        ("Yesterday total spend", yest_total),
        ("Day-over-day delta", today_total - yest_total),
        ("Last 7 day avg (excl. today)", last7_avg),
        ("Last 30 days", mtd_total),
        ("Prior 30 days", last_month_pace),
        ("Δ vs prior 30 days", mtd_total - last_month_pace),
        ("Active orgs today", int((daily[daily["execution_date"] == today][SPEND_COL] > 0).sum()) if not daily.empty else 0),
        ("First-time spenders today", len(first_time)),
        ("Drop-offs (last 7d)", len(drop_offs)),
        ("Top mover up", movers_up.iloc[0]["organization_name"] if not movers_up.empty else "n/a"),
        ("Top mover down", movers_down.iloc[0]["organization_name"] if not movers_down.empty else "n/a"),
    ]

    ws["A1"] = "Daily Execution Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = "$#,##0.00"
        ws.cell(row=i, column=1).fill = SUBTLE_FILL

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def write_table_sheet(ws, title, df, money_cols=None, pct_cols=None, date_cols=None):
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    date_cols = date_cols or []

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if df is None or df.empty:
        ws["A3"] = "No rows."
        return

    cols = list(df.columns)
    for j, c in enumerate(cols, start=1):
        ws.cell(row=3, column=j, value=str(c))
    style_header_row(ws, 3, len(cols))

    for i, row in enumerate(df.itertuples(index=False), start=4):
        for j, val in enumerate(row, start=1):
            colname = cols[j - 1]
            # Blank out $0.00 in money columns so the eye finds the real numbers.
            if (
                colname in money_cols
                and isinstance(val, (int, float, np.floating))
                and not pd.isna(val)
                and abs(float(val)) < 0.005
            ):
                cell = ws.cell(row=i, column=j, value=None)
            else:
                cell = ws.cell(row=i, column=j, value=val if not pd.isna(val) else None)
                if colname in money_cols and isinstance(val, (int, float, np.floating)) and not pd.isna(val):
                    cell.number_format = "$#,##0.00"
                elif colname in pct_cols and isinstance(val, (int, float, np.floating)) and not pd.isna(val):
                    cell.number_format = "0.0\"%\""
                elif colname in date_cols:
                    if isinstance(val, (datetime, pd.Timestamp)):
                        cell.number_format = "yyyy-mm-dd"
            cell.border = BORDER
    autosize(ws, cols)
    ws.freeze_panes = "A4"


def write_pivot_sheet(ws, title, pivot, value_format="$#,##0.00", link_target_sheet: str | None = None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if pivot is None or pivot.empty:
        ws["A3"] = "No rows."
        return

    cols = ["organization_name"] + [
        c.strftime("%Y-%m-%d") if isinstance(c, (datetime, pd.Timestamp)) else str(c) for c in pivot.columns
    ]
    for j, c in enumerate(cols, start=1):
        ws.cell(row=3, column=j, value=c)
    style_header_row(ws, 3, len(cols))

    for i, (org, row) in enumerate(pivot.iterrows(), start=4):
        org_cell = ws.cell(row=i, column=1, value=org)
        org_cell.border = BORDER
        if link_target_sheet:
            # Hyperlink to the Raw sheet so a click jumps the user there to filter
            org_cell.hyperlink = f"#'{link_target_sheet}'!A1"
            org_cell.font = Font(color="1F4E78", underline="single")
        for j, val in enumerate(row.values, start=2):
            # Treat NaN and zero/near-zero as empty so the heatmap and totals
            # aren't drowned in $0.00 cells.
            if val != val or abs(float(val)) < 0.005:
                cell = ws.cell(row=i, column=j, value=None)
            else:
                cell = ws.cell(row=i, column=j, value=float(val))
                cell.number_format = value_format
            cell.border = BORDER

    n_rows = len(pivot)
    n_cols = len(pivot.columns)
    if n_rows >= 1 and n_cols >= 1:
        # Heatmap applies to date columns only — exclude the trailing Total column.
        first_col_letter = get_column_letter(2)
        last_date_col_letter = get_column_letter(1 + n_cols - 1)
        # One Red → Yellow → Green color scale PER ROW so each org is colored
        # against its own min/mid/max (matches =MIN($B3:$AB3) row-locked logic).
        for r in range(4, 4 + n_rows):
            row_range = f"{first_col_letter}{r}:{last_date_col_letter}{r}"
            ws.conditional_formatting.add(
                row_range,
                ColorScaleRule(
                    start_type="min", start_color="F8696B",       # red
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                    end_type="max", end_color="63BE7B",            # green
                ),
            )

    ws.freeze_panes = "B4"
    autosize(ws, cols)


MONEY_METRICS = {"Revenue", "Cost", "Margin", "Avg $/org"}
PCT_METRICS = {"Margin %", "MoM growth %"}
INT_METRICS = {"Executions", "Active orgs"}


def write_monthly_cockpit_sheet(ws, title: str, cockpit: pd.DataFrame) -> None:
    """Render the transposed monthly cockpit: metrics on rows, months on columns."""
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if cockpit is None or cockpit.empty:
        ws["A3"] = "No rows."
        return

    # Header row: blank corner + month labels
    months = list(cockpit.columns)
    ws.cell(row=3, column=1, value="Metric")
    for j, m in enumerate(months, start=2):
        ws.cell(row=3, column=j, value=str(m))
    style_header_row(ws, 3, len(months) + 1)

    metric_names = list(cockpit.index)
    for i, metric in enumerate(metric_names, start=4):
        ws.cell(row=i, column=1, value=str(metric)).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = SUBTLE_FILL
        for j, m in enumerate(months, start=2):
            val = cockpit.loc[metric, m]
            cell = ws.cell(row=i, column=j, value=val if pd.notna(val) else None)
            cell.border = BORDER
            if metric in MONEY_METRICS and pd.notna(val):
                cell.number_format = "$#,##0"
            elif metric in PCT_METRICS and pd.notna(val):
                cell.number_format = "+0.0\"%\";-0.0\"%\";0.0\"%\""
            elif metric in INT_METRICS and pd.notna(val):
                cell.number_format = "#,##0"
            elif str(metric).startswith("# orgs >"):
                cell.number_format = "0"
        ws.cell(row=i, column=1).border = BORDER

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(months) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 14


def write_seedance_usage_sheet(
    ws,
    summary: dict,
    daily: pd.DataFrame,
    pivot: pd.DataFrame,
) -> None:
    """Seedance 2.0 usage over trailing 30 days: summary + daily + workspace pivot."""
    ws["A1"] = "Seedance 2.0 Usage"
    ws["A1"].font = TITLE_FONT

    if not summary or summary.get("total_executions", 0) == 0:
        ws["A3"] = "No Seedance 2.0 usage in the trailing 30-day window."
        return

    window_label = f"{summary['window_start'].strftime('%Y-%m-%d')} → {summary['window_end'].strftime('%Y-%m-%d')}"

    # ---- Section A — Totals ----
    ws["A3"] = f"Trailing {summary['days_in_window']} days: {window_label}"
    ws["A3"].font = Font(bold=True, size=12)

    rows = [
        ("Total executions", summary["total_executions"], "#,##0"),
        ("Total revenue", summary["total_revenue"], "$#,##0.00"),
        ("Total cost", summary["total_cost"], "$#,##0.00"),
        ("Total margin", summary["total_margin"], "$#,##0.00"),
        ("Margin %", summary["margin_pct"], "0.0\"%\""),
        ("Unique workspaces", summary["unique_orgs"], "#,##0"),
        ("Active days (of window)", f"{summary['active_days']} of {summary['days_in_window']}", None),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=5):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = Font(bold=True)
        c1.fill = SUBTLE_FILL
        c1.border = BORDER
        c2 = ws.cell(row=i, column=2, value=value)
        if fmt and isinstance(value, (int, float)):
            c2.number_format = fmt
        c2.border = BORDER

    # ---- Section B — Daily breakdown ----
    row = 5 + len(rows) + 2
    ws.cell(row=row, column=1, value="Daily usage").font = Font(bold=True, size=12)
    row += 1

    daily_cols = ["date", "executions", "revenue", "active workspaces"]
    for j, c in enumerate(daily_cols, start=1):
        ws.cell(row=row, column=j, value=c)
    style_header_row(ws, row, len(daily_cols))
    for r in daily.itertuples(index=False):
        row += 1
        ws.cell(row=row, column=1, value=r.execution_date.strftime("%Y-%m-%d") if hasattr(r.execution_date, "strftime") else r.execution_date).border = BORDER
        c = ws.cell(row=row, column=2, value=int(r.executions))
        c.number_format = "#,##0"; c.border = BORDER
        c = ws.cell(row=row, column=3, value=float(r.revenue))
        c.number_format = "$#,##0.00"; c.border = BORDER
        c = ws.cell(row=row, column=4, value=int(r.active_orgs))
        c.number_format = "0"; c.border = BORDER

    # Heatmap on executions column
    if len(daily) >= 1:
        rng = f"B{row - len(daily) + 1}:B{row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFE699",
                end_type="max", end_color="C65911",
            ),
        )

    # ---- Section C — Workspaces x day pivot ----
    row += 2
    ws.cell(row=row, column=1, value=f"Workspaces × day (executions). All {len(pivot)} workspaces.").font = Font(bold=True, size=12)
    row += 1

    date_cols = [c for c in pivot.columns if c != "Total"]
    header = ["workspace"] + [d.strftime("%m-%d") for d in date_cols] + ["Total"]
    for j, h in enumerate(header, start=1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(header))

    pivot_start_row = row + 1
    for i, (org, row_data) in enumerate(pivot.iterrows(), start=pivot_start_row):
        ws.cell(row=i, column=1, value=str(org)).border = BORDER
        for j, val in enumerate(row_data.values, start=2):
            if val != val or abs(float(val)) < 0.5:
                cell = ws.cell(row=i, column=j, value=None)
            else:
                cell = ws.cell(row=i, column=j, value=int(val))
                cell.number_format = "#,##0"
            cell.border = BORDER

    # Row-wise heatmap on date columns (excluding Total)
    n_rows = len(pivot)
    if n_rows >= 1 and len(date_cols) >= 1:
        first_col_letter = get_column_letter(2)
        last_date_col_letter = get_column_letter(1 + len(date_cols))
        for r in range(pivot_start_row, pivot_start_row + n_rows):
            ws.conditional_formatting.add(
                f"{first_col_letter}{r}:{last_date_col_letter}{r}",
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    mid_type="percentile", mid_value=50, mid_color="FFE699",
                    end_type="max", end_color="C65911",
                ),
            )

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 36


def write_model_usage_sheet(
    ws,
    monthly_winners: pd.DataFrame,
    top_by_exec_30: pd.DataFrame,
    top_by_rev_30: pd.DataFrame,
    window_start,
    window_end,
) -> None:
    """Three sections: monthly winners + top-N (30d) by executions + top-N (30d) by revenue."""
    ws["A1"] = "Model usage"
    ws["A1"].font = TITLE_FONT

    # ---- Section A — Monthly winners ----
    ws["A3"] = f"Monthly model winners (last {MONTHLY_COCKPIT_MONTHS} months)"
    ws["A3"].font = Font(bold=True, size=12)
    row = 4
    if monthly_winners is None or monthly_winners.empty:
        ws.cell(row=row, column=1, value="No data.")
        row += 2
    else:
        cols = list(monthly_winners.columns)
        for j, c in enumerate(cols, start=1):
            ws.cell(row=row, column=j, value=str(c))
        style_header_row(ws, row, len(cols))
        for r in monthly_winners.itertuples(index=False):
            row += 1
            for j, val in enumerate(r, start=1):
                cell = ws.cell(row=row, column=j, value=val)
                colname = cols[j - 1]
                if colname == "executions" and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
                elif colname == "revenue" and isinstance(val, (int, float)):
                    cell.number_format = "$#,##0"
                cell.border = BORDER
        row += 2

    # ---- Section B — Top N by executions in last 30 days ----
    window_label = ""
    if window_start is not pd.NaT and window_end is not pd.NaT:
        window_label = f" — {pd.Timestamp(window_start).date()} → {pd.Timestamp(window_end).date()}"
    ws.cell(row=row, column=1, value=f"Top {MODEL_PIVOT_TOP_N} models by EXECUTIONS, last {MTD_WINDOW_DAYS} days{window_label}").font = Font(bold=True, size=12)
    row += 1
    row = _write_leaderboard(ws, top_by_exec_30, row, primary="executions")
    row += 2

    # ---- Section C — Top N by revenue in last 30 days ----
    ws.cell(row=row, column=1, value=f"Top {MODEL_PIVOT_TOP_N} models by REVENUE, last {MTD_WINDOW_DAYS} days{window_label}").font = Font(bold=True, size=12)
    row += 1
    _write_leaderboard(ws, top_by_rev_30, row, primary="revenue")

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42


def _write_leaderboard(ws, df: pd.DataFrame, start_row: int, primary: str) -> int:
    """Helper: write a leaderboard table (# / model / executions / revenue / margin %)."""
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="No data.")
        return start_row + 1

    cols = list(df.columns)
    for j, c in enumerate(cols, start=1):
        ws.cell(row=start_row, column=j, value=str(c))
    style_header_row(ws, start_row, len(cols))

    for i, r in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(r, start=1):
            colname = cols[j - 1]
            cell = ws.cell(row=i, column=j)
            if colname == "executions" and isinstance(val, (int, float, np.floating)) and pd.notna(val):
                cell.value = float(val)
                cell.number_format = "#,##0"
            elif colname == "revenue" and isinstance(val, (int, float, np.floating)) and pd.notna(val):
                cell.value = float(val)
                cell.number_format = "$#,##0"
            elif colname == "margin %" and isinstance(val, (int, float, np.floating)) and pd.notna(val):
                cell.value = float(val)
                cell.number_format = "0.0\"%\""
            elif colname == "#":
                cell.value = int(val)
                cell.number_format = "0"
            else:
                cell.value = val if not pd.isna(val) else None
            cell.border = BORDER

    # Heatmap on the primary metric column
    if primary in cols:
        col_idx = cols.index(primary) + 1
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}{start_row + 1}:{col_letter}{start_row + len(df)}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFE699",
                end_type="max", end_color="C65911",
            ),
        )

    return start_row + len(df) + 1


def _write_pivot_block(ws, pivot: pd.DataFrame, start_row: int, num_fmt: str) -> int:
    """Helper: write a model-row x week-column pivot starting at start_row.
    Returns the row after the last written row."""
    cols = ["model"] + list(pivot.columns)
    for j, c in enumerate(cols, start=1):
        ws.cell(row=start_row, column=j, value=str(c))
    style_header_row(ws, start_row, len(cols))

    n_rows = len(pivot)
    for i, (model, row_data) in enumerate(pivot.iterrows(), start=start_row + 1):
        ws.cell(row=i, column=1, value=str(model)).border = BORDER
        for j, val in enumerate(row_data.values, start=2):
            if val != val or abs(float(val)) < 0.005:
                cell = ws.cell(row=i, column=j, value=None)
            else:
                cell = ws.cell(row=i, column=j, value=float(val))
                cell.number_format = num_fmt
            cell.border = BORDER

    # Heatmap across the week columns (excluding Total)
    if n_rows >= 1 and len(pivot.columns) > 1:
        first = f"{get_column_letter(2)}{start_row + 1}"
        last_col = 1 + len(pivot.columns) - 1  # exclude Total
        last = f"{get_column_letter(last_col)}{start_row + n_rows}"
        ws.conditional_formatting.add(
            f"{first}:{last}",
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFE699",
                end_type="max", end_color="C65911",
            ),
        )

    return start_row + n_rows + 1


def write_cockpit_sheet(ws, p: dict) -> None:
    """Tab 3 — projection + MoM + tier counts for the current calendar month."""
    ws["A1"] = "Cockpit"
    ws["A1"].font = TITLE_FONT

    if not p:
        ws["A3"] = "No data."
        return

    # Header line: month + elapsed
    ws["A3"] = f"{p['month_label']} — {p['days_elapsed']} of {p['days_in_month']} days elapsed (through {p['through_date']})"
    ws["A3"].font = Font(bold=True, size=12)

    # Projection table — metric x [Actual, Projection (×N), vs prev]
    header = ["", "Actual", f"Projection (×{p['days_in_month']})", f"vs {p['prev_month_label']}"]
    for j, h in enumerate(header, start=1):
        ws.cell(row=5, column=j, value=h)
    style_header_row(ws, 5, len(header))

    rows = [
        ("Revenue", p["rev_actual"], p["rev_proj"], p["mom_rev_pct"], p["prev_rev"]),
        ("Cost", p["cost_actual"], p["cost_proj"], None, None),
        ("Margin", p["margin_actual"], p["margin_proj"], p["mom_margin_pct"], p["prev_margin"]),
        ("Daily avg revenue", p["daily_avg_rev"], None, None, None),
        ("Executions", p["execs_actual"], None, None, None),
    ]
    for i, (label, actual, proj, mom_pct, prev) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = SUBTLE_FILL
        c2 = ws.cell(row=i, column=2, value=actual)
        c2.number_format = "$#,##0" if label != "Executions" else "#,##0"
        if proj is not None:
            c3 = ws.cell(row=i, column=3, value=proj)
            c3.number_format = "$#,##0"
        if mom_pct is not None:
            c4 = ws.cell(row=i, column=4, value=f"{mom_pct:+.1f}%  (was ${prev:,.0f})")
        for j in range(1, 5):
            ws.cell(row=i, column=j).border = BORDER

    # Customer tiers
    tier_header_row = 6 + len(rows) + 2
    ws.cell(row=tier_header_row - 1, column=1, value="Customer tier counts").font = Font(bold=True, size=12)
    th = ["Threshold (>)", "Current (so far)", f"Projected (×{p['days_in_month']}/{p['days_elapsed']})"]
    for j, h in enumerate(th, start=1):
        ws.cell(row=tier_header_row, column=j, value=h)
    style_header_row(ws, tier_header_row, len(th))

    for i, t in enumerate(p["tiers"], start=tier_header_row + 1):
        ws.cell(row=i, column=1, value=t["threshold"]).font = Font(bold=True)
        ws.cell(row=i, column=2, value=t["current"]).number_format = "0"
        ws.cell(row=i, column=3, value=t["projected"]).number_format = "0"
        for j in range(1, 4):
            ws.cell(row=i, column=j).border = BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28


def write_raw_sheet(ws, raw: pd.DataFrame, table_name: str = "RawData") -> None:
    """Write the raw rows as an Excel Table (filterable, sortable)."""
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws["A1"] = "Raw data — filter or sort to drill into any number from the pivot"
    ws["A1"].font = TITLE_FONT

    if raw is None or raw.empty:
        ws["A3"] = "No rows."
        return

    # Restrict to the columns that matter for drill-down; keep order stable.
    keep_cols = [
        c for c in [
            "execution_date",
            "organization_name",
            "model_name",
            "model_category",
            "apikey_owner_email",
            "execution_count",
            "total_revenue",
            "total_provider_cost",
            "total_margin",
        ] if c in raw.columns
    ]
    df = raw[keep_cols].copy().sort_values(["execution_date", "organization_name"], ascending=[False, True])

    header_row = 3
    # Bulk-append via ws.append() instead of per-cell ws.cell() addressing.
    # For ~200k+ row raw dumps, random-access cell writes are dramatically
    # slower than sequential append; this keeps large daily exports fast.
    # Trade-off: per-cell $/date number formats are skipped on this sheet
    # (values are still correct, just displayed as plain numbers/ISO dates).
    ws.append([None] * len(df.columns))  # blank row 2, keep header at row 3
    ws.append([str(c) for c in df.columns])
    date_cols = {i for i, c in enumerate(df.columns) if c == "execution_date"}
    for row in df.itertuples(index=False):
        ws.append([
            (v.strftime("%Y-%m-%d") if i in date_cols and not pd.isna(v) else (None if pd.isna(v) else v))
            for i, v in enumerate(row)
        ])

    last_col = get_column_letter(len(df.columns))
    last_row = header_row + len(df)
    ref = f"A{header_row}:{last_col}{last_row}"

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    # autosize() scans every cell (row*col) to measure width, which is fine for
    # small tables but O(n) prohibitive at 200k+ rows. Use sane fixed widths
    # based on column name/type instead for this sheet.
    for j, c in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        if c in ("organization_name", "model_name", "apikey_owner_email"):
            width = 28
        elif c == "execution_date":
            width = 12
        elif c in ("total_revenue", "total_provider_cost", "total_margin"):
            width = 16
        else:
            width = 16
        ws.column_dimensions[letter].width = width


def build_report(daily: pd.DataFrame, raw: pd.DataFrame, out_path: Path, source_name: str) -> Path:
    movers_up, movers_down, today, yesterday = compute_movers(daily)
    first_time = compute_first_time_today(daily)
    drop_offs = compute_drop_offs(daily)
    mtd_pivot = build_mtd_pivot(daily)
    monthly_pivot = build_monthly_pivot(daily, MONTHLY_PIVOT_MONTHS)
    monthly_cockpit = build_monthly_cockpit(raw, MONTHLY_COCKPIT_MONTHS)
    monthly_model_winners = build_monthly_model_winners(raw, MONTHLY_COCKPIT_MONTHS)
    top_models_exec_30d, top_models_rev_30d, win_start, win_end = build_top_models_30d(raw, MODEL_PIVOT_TOP_N)
    seedance_summary, seedance_daily, seedance_pivot, _, _ = build_seedance_usage(raw)
    projection = build_projection(raw)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    write_summary_sheet(ws, today, yesterday, daily, movers_up, movers_down, first_time, drop_offs)

    # New high-priority cockpit tabs go right after Summary.
    write_cockpit_sheet(wb.create_sheet("Cockpit"), projection)
    write_monthly_cockpit_sheet(
        wb.create_sheet("Monthly cockpit"),
        f"Monthly cockpit (last {MONTHLY_COCKPIT_MONTHS} months)",
        monthly_cockpit,
    )

    write_table_sheet(
        wb.create_sheet("Top movers UP"),
        f"Top movers up — {today.strftime('%Y-%m-%d') if today is not pd.NaT else ''}",
        movers_up.rename(columns={"organization_name": "org"}),
        money_cols=["yesterday", "today", "delta"],
        pct_cols=["pct_change"],
    )
    write_table_sheet(
        wb.create_sheet("Top movers DOWN"),
        f"Top movers down — {today.strftime('%Y-%m-%d') if today is not pd.NaT else ''}",
        movers_down.rename(columns={"organization_name": "org"}),
        money_cols=["yesterday", "today", "delta"],
        pct_cols=["pct_change"],
    )
    write_table_sheet(
        wb.create_sheet("First-time today"),
        "First-time spenders today",
        first_time.rename(columns={"organization_name": "org"}),
        money_cols=["today_spend"],
        date_cols=["first_spend_date"],
    )
    write_table_sheet(
        wb.create_sheet("Drop-offs"),
        f"Drop-offs (spent in last {DROP_OFF_LOOKBACK_DAYS}d, silent today)",
        drop_offs.rename(columns={"organization_name": "org"}),
        money_cols=["lookback_total"],
        date_cols=["last_spend_date"],
    )
    write_pivot_sheet(
        wb.create_sheet("MTD pivot"),
        f"Trailing {MTD_WINDOW_DAYS}-day org spend",
        mtd_pivot,
    )
    write_pivot_sheet(
        wb.create_sheet("Monthly pivot"),
        f"Monthly org spend (last {MONTHLY_PIVOT_MONTHS} months)",
        monthly_pivot,
    )

    write_model_usage_sheet(
        wb.create_sheet("Model usage"),
        monthly_model_winners,
        top_models_exec_30d,
        top_models_rev_30d,
        win_start,
        win_end,
    )

    write_seedance_usage_sheet(
        wb.create_sheet("Seedance 2.0 Usage"),
        seedance_summary,
        seedance_daily,
        seedance_pivot,
    )

    write_raw_sheet(wb.create_sheet("Raw"), raw)

    # Source provenance sheet
    prov = wb.create_sheet("Source")
    prov["A1"] = "Source provenance"
    prov["A1"].font = TITLE_FONT
    prov["A3"] = "Source file"
    prov["B3"] = source_name
    prov["A4"] = "Generated at"
    prov["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    prov["A5"] = "Rows in raw data"
    prov["B5"] = len(raw)
    prov["A6"] = "Date range"
    if not raw.empty:
        prov["B6"] = f"{raw['execution_date'].min().date()} -> {raw['execution_date'].max().date()}"

    wb.save(out_path)
    return out_path


def list_all_inbox_files(inbox: Path) -> list[Path]:
    """Return every supported source file in inbox, sorted oldest → newest by
    the date embedded in the filename when present (falls back to mtime)."""
    candidates: list[Path] = []
    for ext in SUPPORTED_EXTS:
        candidates.extend(inbox.glob(f"execution_report_*{ext}"))
    candidates = [p for p in candidates if not p.name.startswith("~$")]

    def sort_key(p: Path):
        # Try to extract YYYY-MM-DD from filename for stable chronological order
        import re as _re
        m = _re.search(r"(\d{4}-\d{2}-\d{2})", p.name)
        return (m.group(1) if m else "", p.stat().st_mtime)

    return sorted(candidates, key=sort_key)


def process_one(src: Path, reports: Path, archive: Path) -> dict:
    """Generate the daily report for one source file and archive the source.
    Returns a summary dict."""
    raw = load_raw(src)
    daily = org_daily_spend(raw)
    today = daily["execution_date"].max()
    today_str = today.strftime("%Y-%m-%d") if today is not pd.NaT else "unknown"
    out_path = reports / f"eachlabs daily report {today_str}.xlsx"
    build_report(daily, raw, out_path, source_name=src.name)

    archived = archive / src.name
    if archived.exists():
        ts = datetime.now().strftime("%H%M%S")
        archived = archive / f"{src.stem}_{ts}{src.suffix}"
    try:
        shutil.move(str(src), str(archived))
    except Exception:
        shutil.copy2(str(src), str(archived))
        try:
            src.unlink()
        except Exception:
            pass

    return {
        "source": str(archived),
        "report": str(out_path),
        "report_date": today_str,
        "orgs_active_today": int((daily[daily["execution_date"] == today][SPEND_COL] > 0).sum()),
        "today_total_spend": float(daily[daily["execution_date"] == today][SPEND_COL].sum()),
    }


def main(base_dir: Path):
    inbox = base_dir / "inbox"
    archive = base_dir / "archive"
    reports = base_dir / "eachlabs daily report"
    state_dir = base_dir / "state"
    for p in [inbox, archive, reports, state_dir]:
        p.mkdir(parents=True, exist_ok=True)

    files = list_all_inbox_files(inbox)
    if not files:
        msg = {
            "status": "no_file",
            "message": f"No execution_report_*.(xlsx|csv) in {inbox}",
        }
        print(json.dumps(msg))
        sys.exit(2)

    results: list[dict] = []
    errors: list[dict] = []
    for src in files:
        try:
            r = process_one(src, reports, archive)
            results.append(r)
        except Exception as e:
            errors.append({"source": src.name, "error": f"{type(e).__name__}: {e}"})

    summary = {
        "status": "ok" if results else "error",
        "processed_count": len(results),
        "results": results,
        "errors": errors,
    }
    (state_dir / "last_run.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))
    if errors and not results:
        sys.exit(4)


def build_report_from_csv(csv_path: Path, out_dir: Path) -> tuple[Path, str]:
    """Cloud-friendly entry point. Reads a single CSV, builds the report xlsx,
    writes it to out_dir, returns (output_path, report_date_string).

    No inbox/archive/state side effects. Safe to call in a temp directory.
    """
    raw = load_raw(csv_path)
    daily = org_daily_spend(raw)
    today = daily["execution_date"].max()
    report_date = today.strftime("%Y-%m-%d") if today is not pd.NaT else "unknown"
    out_path = out_dir / f"eachlabs daily report {report_date}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_report(daily, raw, out_path, source_name=csv_path.name)
    return out_path, report_date


if __name__ == "__main__":
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent
    main(base)
